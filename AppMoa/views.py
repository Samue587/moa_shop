import os
import uuid
from decimal import Decimal
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.conf import settings
from django.db.models import Sum, Count, F, ExpressionWrapper, FloatField
from django.db.models.functions import TruncDay
from AppMoa.decorators import permiso_requerido

import resend
from .models import TokenReset


from .models import (
    Rol, Permiso, RolPermiso, Usuario,
    Proveedor, Catalogo, Categoria, Producto, VariacionProducto,
    Entrada, DetalleEntrada,
    Venta, DetalleVenta,
    Envio,
)

#=================PERMISOS=========================================

def permisos_usuario(request):

    permisos = []

    if request.user.is_authenticated:

        permisos = request.user.rol.permisos.values_list(
            'slug',
            flat=True
        )

    return {
        'permisos_usuario': permisos
    }

UPLOAD_DIR = settings.MEDIA_ROOT

CIUDADES_POR_DEPARTAMENTO = {
    'Antioquia':          ['Medellín', 'Bello', 'Itagüí', 'Envigado', 'Rionegro'],
    'Atlántico':          ['Barranquilla', 'Soledad', 'Malambo'],
    'Bogotá D.C.':        ['Bogotá'],
    'Bolívar':            ['Cartagena', 'Magangué', 'Turbaco'],
    'Boyacá':             ['Tunja', 'Duitama', 'Sogamoso'],
    'Caldas':             ['Manizales', 'La Dorada', 'Chinchiná'],
    'Cauca':              ['Popayán', 'Santander de Quilichao'],
    'Cesar':              ['Valledupar', 'Aguachica'],
    'Córdoba':            ['Montería', 'Lorica', 'Cereté'],
    'Cundinamarca':       ['Soacha', 'Fusagasugá', 'Zipaquirá', 'Facatativá'],
    'Huila':              ['Neiva', 'Pitalito', 'Garzón'],
    'La Guajira':         ['Riohacha', 'Maicao'],
    'Magdalena':          ['Santa Marta', 'Ciénaga'],
    'Meta':               ['Villavicencio', 'Acacías'],
    'Nariño':             ['Pasto', 'Tumaco', 'Ipiales'],
    'Norte de Santander': ['Cúcuta', 'Ocaña', 'Pamplona'],
    'Quindío':            ['Armenia', 'Calarcá'],
    'Risaralda':          ['Pereira', 'Dosquebradas'],
    'Santander':          ['Bucaramanga', 'Floridablanca', 'Barrancabermeja'],
    'Sucre':              ['Sincelejo', 'Corozal'],
    'Tolima':             ['Ibagué', 'Espinal', 'Honda'],
    'Valle del Cauca':    ['Cali', 'Buenaventura', 'Palmira', 'Tuluá'],
    'Putumayo':           ['Mocoa', 'Puerto Asís', 'Orito', 'La Hormiga']
}


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                         DECORADORES                                  ║
# ╚══════════════════════════════════════════════════════════════════════╝

def login_required_custom(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('usuario_id'):
            return redirect('login')
        try:
            usuario = Usuario.objects.get(pk=request.session['usuario_id'])
            if not usuario.activo:
                request.session.flush()
                messages.error(request, 'Tu cuenta ha sido desactivada.')
                return redirect('login')
        except Usuario.DoesNotExist:
            request.session.flush()
            return redirect('login')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        uid = request.session.get('usuario_id')
        if not uid:
            return redirect('login')
        try:
            usuario = Usuario.objects.select_related('rol').get(pk=uid)
            if not usuario.activo:
                request.session.flush()
                messages.error(request, 'Tu cuenta ha sido desactivada.')
                return redirect('login')
            if not usuario.is_admin:
                return redirect('tienda')
        except Usuario.DoesNotExist:
            request.session.flush()
            return redirect('login')
        request.usuario = usuario
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def get_usuario_sesion(request):
    uid = request.session.get('usuario_id')
    if not uid:
        return None
    try:
        usuario = Usuario.objects.select_related('rol').get(pk=uid)
        if not usuario.activo:
            request.session.flush()
            return None
        return usuario
    except Usuario.DoesNotExist:
        return None


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                        AUTENTICACIÓN                                 ║
# ╚══════════════════════════════════════════════════════════════════════╝

def home(request):
    return redirect('tienda')


def login_view(request):
    if request.session.get('usuario_id'):
        u = get_usuario_sesion(request)
        return redirect('admin_dashboard' if u and u.is_admin else 'tienda')

    if request.method == 'POST':
        email    = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        try:
            usuario = Usuario.objects.select_related('rol').get(correo_usuario=email)
        except Usuario.DoesNotExist:
            return render(request, 'login.html', {'error': 'Credenciales incorrectas.'})

        if not check_password(password, usuario.contrasena):
            return render(request, 'login.html', {'error': 'Credenciales incorrectas.'})

        if not usuario.activo:
            return render(request, 'login.html', {'error': 'Tu cuenta está inactiva. Contacta al administrador.'})

        request.session['usuario_id']     = usuario.id
        request.session['usuario_nombre'] = usuario.get_full_name()
        request.session['usuario_rol']    = usuario.rol.nombre_rol if usuario.rol else ''
        request.session.save()

        return redirect('admin_dashboard' if usuario.is_admin else 'tienda')

    return render(request, 'login.html')


def logout_view(request):
    request.session.flush()
    return redirect('tienda')


def registro_view(request):
    if request.method == 'POST':
        nombres   = request.POST.get('nombre', '').strip()
        apellidos = request.POST.get('apellido', '').strip()
        email     = request.POST.get('email', '').strip()
        password  = request.POST.get('password', '')

        if not nombres or not email or not password:
            return render(request, 'registro.html', {'error': 'Todos los campos son obligatorios.'})

        if Usuario.objects.filter(correo_usuario=email).exists():
            return render(request, 'registro.html', {'error': 'El email ya está registrado.'})

        try:
            rol_cliente = Rol.objects.get(nombre_rol='cliente')
        except Rol.DoesNotExist:
            return render(request, 'registro.html', {'error': 'Error del sistema. Contacte al administrador.'})

        u = Usuario(
            nombres_usuario=nombres,
            apellidos_usuario=apellidos,
            correo_usuario=email,
            rol=rol_cliente
        )
        u.set_password(password)
        u.save()

        messages.success(request, 'Registro exitoso. Ya puedes iniciar sesión.')
        return redirect('login')

    return render(request, 'registro.html')


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                        TIENDA PÚBLICA                                ║
# ╚══════════════════════════════════════════════════════════════════════╝

def tienda(request):
    categoria_id = request.GET.get('categoria')
    catalogo_id  = request.GET.get('catalogo')
    busqueda     = request.GET.get('q', '').strip()

    productos = Producto.objects.select_related('categoria', 'catalogo', 'proveedor').filter(estado='DISPONIBLE')
    categorias = Categoria.objects.filter(estado='ACTIVO')
    catalogos  = Catalogo.objects.all()

    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    if catalogo_id:
        productos = productos.filter(catalogo_id=catalogo_id)
    if busqueda:
        productos = productos.filter(nombre_producto__icontains=busqueda)

    return render(request, 'tienda/index.html', {
        'productos':   productos,
        'categorias':  categorias,
        'catalogos':   catalogos,
        'busqueda':    busqueda,
        'usuario':     get_usuario_sesion(request),
        'cantidad_carrito': _cantidad_carrito(request),
        'MEDIA_URL':        settings.MEDIA_URL,
    })


def detalle_producto(request, id):
    producto   = get_object_or_404(Producto.objects.select_related('categoria', 'proveedor'), pk=id)
    variaciones = VariacionProducto.objects.filter(producto=producto, stock_actual__gt=0)
    return render(request, 'tienda/detalle.html', {
        'producto':         producto,
        'variaciones':      variaciones,
        'usuario':          get_usuario_sesion(request),
        'cantidad_carrito': _cantidad_carrito(request),
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                           CARRITO                                    ║
# ╚══════════════════════════════════════════════════════════════════════╝

def _get_carrito(request):
    if 'carrito' not in request.session:
        request.session['carrito'] = {'items': []}
        request.session.modified = True
    return request.session['carrito']


def _save_carrito(request, carrito):
    request.session['carrito'] = carrito
    request.session.modified   = True
    request.session.save()


def _calcular_total(carrito):
    return sum(Decimal(str(i['subtotal'])) for i in carrito['items'])


def _cantidad_carrito(request):
    carrito = request.session.get('carrito', {'items': []})
    return sum(i['cantidad'] for i in carrito['items'])


def carrito_ver(request):
    usuario = get_usuario_sesion(request)
    if not usuario:
        messages.error(request, 'Debes iniciar sesión para ver tu carrito.')
        return redirect('login')

    carrito = _get_carrito(request)
    return render(request, 'tienda/carrito.html', {
        'carrito':          carrito,
        'items':            carrito['items'],
        'total':            _calcular_total(carrito),
        'usuario':          usuario,
        'cantidad_carrito': _cantidad_carrito(request),
    })


def carrito_agregar(request):
    if not request.session.get('usuario_id'):
        messages.error(request, 'Debes iniciar sesión para agregar productos.')
        return redirect('login')

    if request.method != 'POST':
        return redirect('tienda')

    variacion_id = request.POST.get('variacion_id')
    cantidad     = request.POST.get('cantidad', 1)

    if not variacion_id:
        messages.error(request, 'Variación no válida.')
        return redirect('tienda')

    try:
        cantidad     = int(cantidad)
        variacion_id = int(variacion_id)
    except (ValueError, TypeError):
        messages.error(request, 'Datos inválidos.')
        return redirect('tienda')

    if cantidad <= 0:
        messages.error(request, 'La cantidad debe ser mayor a 0.')
        return redirect('tienda')

    try:
        variacion = VariacionProducto.objects.select_related('producto').get(pk=variacion_id)
    except VariacionProducto.DoesNotExist:
        messages.error(request, 'Variación no encontrada.')
        return redirect('tienda')

    if variacion.sin_stock:
        messages.error(request, 'Esta variación no tiene stock disponible.')
        return redirect('tienda')

    carrito        = _get_carrito(request)
    item_existente = next((i for i in carrito['items'] if i['variacion_id'] == variacion.id), None)
    cant_actual    = item_existente['cantidad'] if item_existente else 0

    if variacion.stock_actual < (cant_actual + cantidad):
        messages.error(request, f'Stock insuficiente. Disponibles: {variacion.stock_actual}.')
        return redirect('tienda')

    precio = variacion.precio_con_iva

    if item_existente:
        item_existente['cantidad'] += cantidad
        item_existente['subtotal']  = round(precio * item_existente['cantidad'], 2)
    else:
        carrito['items'].append({
            'variacion_id':   variacion.id,
            'producto_id':    variacion.producto.id,
            'nombre':         variacion.producto.nombre_producto,
            'talla':          variacion.talla,
            'color':          variacion.color,
            'sku':            variacion.sku_unico,
            'precio':         round(precio, 2),
            'foto':           variacion.producto.foto or '',
            'cantidad':       cantidad,
            'subtotal':       round(precio * cantidad, 2),
        })

    _save_carrito(request, carrito)
    messages.success(request, f'"{variacion.producto.nombre_producto}" agregado al carrito.')
    return redirect('carrito_ver')


def carrito_actualizar(request, variacion_id):
    if not request.session.get('usuario_id'):
        return redirect('login')

    if request.method != 'POST':
        return redirect('carrito_ver')

    try:
        cantidad = int(request.POST.get('cantidad', 1))
    except (ValueError, TypeError):
        messages.error(request, 'Cantidad inválida.')
        return redirect('carrito_ver')

    if cantidad <= 0:
        messages.error(request, 'La cantidad debe ser mayor a 0.')
        return redirect('carrito_ver')

    try:
        variacion = VariacionProducto.objects.get(pk=variacion_id)
    except VariacionProducto.DoesNotExist:
        messages.error(request, 'Variación no encontrada.')
        return redirect('carrito_ver')

    if variacion.stock_actual < cantidad:
        messages.error(request, f'Stock insuficiente. Disponibles: {variacion.stock_actual}.')
        return redirect('carrito_ver')

    carrito = _get_carrito(request)
    for item in carrito['items']:
        if item['variacion_id'] == int(variacion_id):
            item['cantidad'] = cantidad
            item['subtotal'] = round(item['precio'] * cantidad, 2)
            break

    _save_carrito(request, carrito)
    messages.success(request, 'Cantidad actualizada.')
    return redirect('carrito_ver')


def carrito_remover(request, variacion_id):
    if not request.session.get('usuario_id'):
        return redirect('login')

    if request.method != 'POST':
        return redirect('carrito_ver')

    carrito = _get_carrito(request)
    carrito['items'] = [i for i in carrito['items'] if i['variacion_id'] != int(variacion_id)]
    _save_carrito(request, carrito)
    messages.success(request, 'Producto eliminado del carrito.')
    return redirect('carrito_ver')


def carrito_limpiar(request):
    if not request.session.get('usuario_id'):
        return redirect('login')

    if request.method == 'POST':
        _save_carrito(request, {'items': []})
        messages.success(request, 'Carrito vaciado.')
    return redirect('carrito_ver')


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                          CHECKOUT                                    ║
# ╚══════════════════════════════════════════════════════════════════════╝

@login_required_custom
def checkout_procesar(request):
    usuario = get_usuario_sesion(request)
    carrito = _get_carrito(request)
    list(messages.get_messages(request))
    storage = messages.get_messages(request)
    storage.used = True
    

    if not carrito['items']:
        messages.error(request, 'El carrito está vacío.')
        return redirect('carrito_ver')

    ctx_base = {
        'usuario':                   usuario,
        'items':                     carrito['items'],
        'total':                     _calcular_total(carrito),
        'cantidad_carrito':          _cantidad_carrito(request),
        'departamentos':             sorted(CIUDADES_POR_DEPARTAMENTO.keys()),
        'ciudades_por_departamento': CIUDADES_POR_DEPARTAMENTO,
    }

    if request.method == 'GET':
        return render(request, 'tienda/checkout.html', {**ctx_base, 'form': {}, 'errores': []})

    direccion    = request.POST.get('direccion', '').strip()
    barrio       = request.POST.get('barrio', '').strip()
    ciudad       = request.POST.get('ciudad', '').strip()
    departamento = request.POST.get('departamento', '').strip()
    tipo_vivienda = request.POST.get('tipo_vivienda', 'CASA').strip()
    especificaciones = request.POST.get('especificaciones', '').strip()
    telefono     = request.POST.get('telefono', '').strip()
    empresa      = request.POST.get('empresa', 'Interrapidísimo').strip()

    errores = []
    if not direccion:    errores.append('La dirección es obligatoria.')
    if not ciudad:       errores.append('La ciudad es obligatoria.')
    if not departamento: errores.append('El departamento es obligatorio.')
    if not telefono:     errores.append('El teléfono es obligatorio.')

    if errores:
        return render(request, 'tienda/checkout.html', {
            **ctx_base,
            'form': request.POST,
            'errores': errores,
        })

    try:
        with transaction.atomic():
            # Calcular totales
            total    = float(_calcular_total(carrito))
            iva      = round(total * 0.19 / 1.19, 2)
            subtotal = round(total - iva, 2)

            venta = Venta.objects.create(
                usuario=usuario,
                subtotal=subtotal,
                monto_iva=iva,
                monto_final=total,
            )

            for item in carrito['items']:
                variacion = VariacionProducto.objects.select_for_update().get(pk=item['variacion_id'])
                if variacion.stock_actual < item['cantidad']:
                    raise Exception(f"Stock insuficiente para '{variacion.producto.nombre_producto}' ({variacion.talla} - {variacion.color}).")

                DetalleVenta.objects.create(
                    venta=venta,
                    variacion=variacion,
                    cantidad=item['cantidad'],
                    precio_unitario=item['precio'],
                )

                variacion.stock_actual -= item['cantidad']
                variacion.save(update_fields=['stock_actual', 'actualizado_en'])

                producto = variacion.producto
                producto.actualizar_estado()
                producto.save(update_fields=['estado', 'actualizado_en'])

            Envio.objects.create(
                usuario=usuario,
                venta=venta,
                departamento_envio=departamento,
                ciudad_envio=ciudad,
                barrio_envio=barrio or None,
                direccion_envio=direccion,
                tipo_vivienda=tipo_vivienda,
                especificaciones_llegada=especificaciones or None,
                telefono_llegada=telefono,
                empresa_transportadora='SIN ASIGNAR TRANSPORTISTA',
                estado_envio='PENDIENTE',
                fecha_envio=date.today(),
                fecha_estipulada_llegada=date.today() + timedelta(days=3),
                costo_envio=Decimal('0.00'),
            )

            _save_carrito(request, {'items': []})

        messages.success(request, f'¡Compra realizada con éxito! Orden #{venta.id}')
        return redirect('checkout_comprobante', venta_id=venta.id)

    except Exception as e:
        messages.error(request, f'Error al procesar la compra: {e}')
        return render(request, 'tienda/checkout.html', {**ctx_base, 'form': request.POST})


def checkout_comprobante(request, venta_id):
    usuario = get_usuario_sesion(request)
    if not usuario:
        return redirect('login')

    venta    = get_object_or_404(Venta.objects.select_related('usuario'), pk=venta_id)
    detalles = DetalleVenta.objects.select_related('variacion__producto').filter(venta=venta)
    envio    = Envio.objects.filter(venta=venta).first()

    return render(request, 'tienda/comprobante.html', {
        'venta':    venta,
        'detalles': detalles,
        'envio':    envio,
        'usuario':  usuario,
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                        PERFIL USUARIO                                ║
# ╚══════════════════════════════════════════════════════════════════════╝

@login_required_custom
def perfil_usuario(request):
    usuario = get_usuario_sesion(request)

    ventas = Venta.objects.filter(usuario=usuario).prefetch_related(
        'detalles__variacion__producto',
        'envio'
    ).order_by('-id')

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'actualizar_datos':
            nombres   = request.POST.get('nombres', '').strip()
            apellidos = request.POST.get('apellidos', '').strip()
            email     = request.POST.get('email', '').strip()

            if not nombres or not email:
                messages.error(request, 'Nombre y email son obligatorios.')
                return render(request, 'tienda/perfil.html', {'usuario': usuario, 'ventas': ventas})

            if Usuario.objects.filter(correo_usuario=email).exclude(pk=usuario.id).exists():
                messages.error(request, 'Ese email ya está en uso.')
                return render(request, 'tienda/perfil.html', {'usuario': usuario, 'ventas': ventas})

            usuario.nombres_usuario   = nombres
            usuario.apellidos_usuario = apellidos
            usuario.correo_usuario    = email
            usuario.save()
            request.session['usuario_nombre'] = usuario.get_full_name()
            messages.success(request, 'Datos actualizados correctamente.')
            return redirect('perfil_usuario')

        elif accion == 'cambiar_password':
            password_actual = request.POST.get('password_actual', '')
            password_nuevo  = request.POST.get('password_nuevo', '')
            password_conf   = request.POST.get('password_confirmacion', '')

            if not check_password(password_actual, usuario.contrasena):
                messages.error(request, 'La contraseña actual es incorrecta.')
                return render(request, 'tienda/perfil.html', {'usuario': usuario, 'ventas': ventas})

            if len(password_nuevo) < 6:
                messages.error(request, 'La nueva contraseña debe tener al menos 6 caracteres.')
                return render(request, 'tienda/perfil.html', {'usuario': usuario, 'ventas': ventas})

            if password_nuevo != password_conf:
                messages.error(request, 'Las contraseñas no coinciden.')
                return render(request, 'tienda/perfil.html', {'usuario': usuario, 'ventas': ventas})

            usuario.set_password(password_nuevo)
            usuario.save()
            messages.success(request, 'Contraseña cambiada correctamente.')
            return redirect('perfil_usuario')

    return render(request, 'tienda/perfil.html', {
        'usuario':          usuario,
        'ventas':           ventas,
        'total_compras':    ventas.count(),
        'total_gastado':    sum(v.total_venta for v in ventas),
        'cantidad_carrito': _cantidad_carrito(request),
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                      DASHBOARD ADMIN                                 ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
def panel_admin(request):
    context = {
        'total_productos':     Producto.objects.count(),
        'total_categorias':    Categoria.objects.count(),
        'total_usuarios':      Usuario.objects.count(),
        'total_roles':         Rol.objects.count(),
        'total_ventas':        Venta.objects.count(),
        'total_proveedores':   Proveedor.objects.count(),
        'productos_agotados':  Producto.objects.filter(estado='NO_DISPONIBLE').count(),
        'ventas_hoy':          Venta.objects.filter(fecha__date=timezone.now().date()).count(),
        'ingresos_hoy':        Venta.objects.filter(
                                   fecha__date=timezone.now().date()
                               ).aggregate(total=Sum('monto_final'))['total'] or 0,
        'productos_recientes': Producto.objects.select_related('categoria').order_by('-creado_en')[:8],
        'ventas_recientes':    Venta.objects.select_related('usuario').order_by('-fecha')[:5],
        'envios_pendientes':   Envio.objects.exclude(estado_envio__in=['ENTREGADO', 'CANCELADO']).count(),
        'stock_bajo':          VariacionProducto.objects.filter(
                                   stock_actual__lte=models_stock_minimo()
                               ).count() if False else
                               sum(1 for v in VariacionProducto.objects.all() if v.bajo_stock),
        'usuario':             request.usuario,
    }
    return render(request, 'admin/dashboard.html', context)


@admin_required
def dashboard_stats(request):
    return JsonResponse({
        'ventas_hoy':        Venta.objects.filter(fecha__date=timezone.now().date()).count(),
        'total_ventas':      Venta.objects.count(),
        'envios_pendientes': Envio.objects.exclude(estado_envio__in=['ENTREGADO', 'CANCELADO']).count(),
    })


def models_stock_minimo():
    """Helper para evitar import circular en anotación."""
    return 0


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                           ROLES                                      ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_roles')
def listar_roles(request):
    roles = Rol.objects.prefetch_related('rol_permisos__permiso').order_by('id')
    return render(request, 'admin/roles/lista.html', {
        'roles': roles, 'usuario': request.usuario
    })
    
def asignar_permisos():

    admin = Rol.objects.get(
        nombre_rol='Administrador'
    )

    admin.permisos.set(
        Permiso.objects.all()
    )

   

@admin_required
@permiso_requerido('crear_roles')
def crear_rol(request):
    permisos = Permiso.objects.all().order_by('slug')
    if request.method == 'POST':
        nombre_rol        = request.POST.get('nombre_rol', '').strip().upper()
        permisos_ids      = request.POST.getlist('permisos')

        if not nombre_rol:
            messages.error(request, 'El nombre es obligatorio.')
            return render(request, 'admin/roles/crear.html', {'form': request.POST, 'permisos': permisos, 'usuario': request.usuario})

        if Rol.objects.filter(nombre_rol__iexact=nombre_rol).exists():
            messages.error(request, f'Ya existe un rol "{nombre_rol}".')
            return render(request, 'admin/roles/crear.html', {'form': request.POST, 'permisos': permisos, 'usuario': request.usuario})

        with transaction.atomic():
            rol = Rol.objects.create(nombre_rol=nombre_rol)
            for pid in permisos_ids:
                try:
                    RolPermiso.objects.create(rol=rol, permiso=Permiso.objects.get(pk=pid))
                except Permiso.DoesNotExist:
                    pass

        messages.success(request, f'Rol "{nombre_rol}" creado.')
        return redirect('admin_roles')

    return render(request, 'admin/roles/crear.html', {'form': {}, 'permisos': permisos, 'usuario': request.usuario})


@admin_required
@permiso_requerido('editar_roles')
def editar_rol(request, id):
    rol      = get_object_or_404(Rol, pk=id)
    permisos = Permiso.objects.all().order_by('slug')
    permisos_del_rol = RolPermiso.objects.filter(rol=rol).values_list('permiso_id', flat=True)

    if request.method == 'POST':
        nombre_rol   = request.POST.get('nombre_rol', '').strip().upper()
        permisos_ids = request.POST.getlist('permisos')

        if not nombre_rol:
            messages.error(request, 'El nombre es obligatorio.')
            return render(request, 'admin/roles/editar.html', {
                'rol': rol, 'permisos': permisos,
                'permisos_del_rol': permisos_del_rol,
                'form': request.POST, 'usuario': request.usuario
            })

        if Rol.objects.filter(nombre_rol__iexact=nombre_rol).exclude(pk=id).exists():
            messages.error(request, 'Ya existe otro rol con ese nombre.')
            return render(request, 'admin/roles/editar.html', {
                'rol': rol, 'permisos': permisos,
                'permisos_del_rol': permisos_del_rol,
                'form': request.POST, 'usuario': request.usuario
            })

        with transaction.atomic():
            rol.nombre_rol = nombre_rol
            rol.save()
            RolPermiso.objects.filter(rol=rol).delete()
            for pid in permisos_ids:
                try:
                    RolPermiso.objects.create(rol=rol, permiso=Permiso.objects.get(pk=pid))
                except Permiso.DoesNotExist:
                    pass

        messages.success(request, f'Rol "{nombre_rol}" actualizado.')
        return redirect('admin_roles')  

    return render(request, 'admin/roles/editar.html', {
        'rol': rol, 'permisos': permisos,
        'permisos_del_rol': list(permisos_del_rol),
        'form': {}, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('eliminar_roles')
def eliminar_rol(request, id):
    rol            = get_object_or_404(Rol, pk=id)
    usuarios_count = Usuario.objects.filter(rol=rol).count()

    if request.method == 'POST':
        nombre = rol.nombre_rol
        rol.delete()
        messages.success(request, f'Rol "{nombre}" eliminado.')
        return redirect('admin_roles')

    return render(request, 'admin/roles/eliminar.html', {
        'rol': rol, 'usuarios_count': usuarios_count, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                          PERMISOS                                    ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
def listar_permisos(request):
    permisos = Permiso.objects.all().order_by('slug')
    return render(request, 'admin/permisos/lista.html', {
        'permisos': permisos, 'usuario': request.usuario
    })


@admin_required
def crear_permiso(request):
    if request.method == 'POST':
        descripcion = request.POST.get('descripcion', '').strip()
        slug        = request.POST.get('slug', '').strip().lower().replace(' ', '-')

        if not descripcion or not slug:
            messages.error(request, 'Descripción y slug son obligatorios.')
            return render(request, 'admin/permisos/crear.html', {'form': request.POST, 'usuario': request.usuario})

        if Permiso.objects.filter(slug=slug).exists():
            messages.error(request, f'Ya existe un permiso con slug "{slug}".')
            return render(request, 'admin/permisos/crear.html', {'form': request.POST, 'usuario': request.usuario})

        Permiso.objects.create(descripcion=descripcion, slug=slug)
        messages.success(request, f'Permiso "{slug}" creado.')
        return redirect('admin_permisos')

    return render(request, 'admin/permisos/crear.html', {'form': {}, 'usuario': request.usuario})


@admin_required
def editar_permiso(request, id):
    permiso = get_object_or_404(Permiso, pk=id)

    if request.method == 'POST':
        descripcion = request.POST.get('descripcion', '').strip()
        slug        = request.POST.get('slug', '').strip().lower().replace(' ', '-')

        if not descripcion or not slug:
            messages.error(request, 'Descripción y slug son obligatorios.')
            return render(request, 'admin/permisos/editar.html', {
                'permiso': permiso, 'form': request.POST, 'usuario': request.usuario
            })

        if Permiso.objects.filter(slug=slug).exclude(pk=id).exists():
            messages.error(request, 'Ya existe otro permiso con ese slug.')
            return render(request, 'admin/permisos/editar.html', {
                'permiso': permiso, 'form': request.POST, 'usuario': request.usuario
            })

        permiso.descripcion = descripcion
        permiso.slug        = slug
        permiso.save()
        messages.success(request, f'Permiso "{slug}" actualizado.')
        return redirect('admin_permisos')

    return render(request, 'admin/permisos/editar.html', {
        'permiso': permiso, 'form': {}, 'usuario': request.usuario
    })


@admin_required
def eliminar_permiso(request, id):
    permiso = get_object_or_404(Permiso, pk=id)
    if request.method == 'POST':
        slug = permiso.slug
        permiso.delete()
        messages.success(request, f'Permiso "{slug}" eliminado.')
        return redirect('admin_permisos')
    return render(request, 'admin/permisos/eliminar.html', {
        'permiso': permiso, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                          USUARIOS                                    ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_usuarios')
def listar_usuarios(request):
    busqueda = request.GET.get('q', '').strip()
    usuarios = Usuario.objects.select_related('rol').order_by('-creado_en')
    if busqueda:
        usuarios = usuarios.filter(
            Q(nombres_usuario__icontains=busqueda) |
            Q(apellidos_usuario__icontains=busqueda) |
            Q(correo_usuario__icontains=busqueda)
        )
    return render(request, 'admin/usuarios/lista.html', {
        'usuarios': usuarios, 'busqueda': busqueda, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('crear_usuarios')
def crear_usuario(request):
    roles = Rol.objects.all()
    if request.method == 'POST':
        nombres          = request.POST.get('nombres', '').strip()
        apellidos        = request.POST.get('apellidos', '').strip()
        email            = request.POST.get('email', '').strip()
        password         = request.POST.get('password', '')
        rol_id           = request.POST.get('rol')
        estado           = request.POST.get('estado', 'ACTIVO')
        tipo_documento   = request.POST.get('tipo_documento', '')
        numero_documento = request.POST.get('numero_documento', '').strip()

        if not nombres or not email or not password or not rol_id:
            messages.error(request, 'Nombre, email, contraseña y rol son obligatorios.')
            return render(request, 'admin/usuarios/crear.html', {
                'form': request.POST, 'roles': roles, 'usuario': request.usuario
            })

        if Usuario.objects.filter(correo_usuario=email).exists():
            messages.error(request, f'El email "{email}" ya está registrado.')
            return render(request, 'admin/usuarios/crear.html', {
                'form': request.POST, 'roles': roles, 'usuario': request.usuario
            })

        u = Usuario(
            nombres_usuario=nombres,
            apellidos_usuario=apellidos,
            correo_usuario=email,
            rol=get_object_or_404(Rol, pk=rol_id),
            estado_usuario=estado,
            tipo_documento=tipo_documento or None,
            numero_documento=numero_documento or None,
        )
        u.set_password(password)
        u.save()
        messages.success(request, f'Usuario "{email}" creado.')
        return redirect('admin_usuarios')

    return render(request, 'admin/usuarios/crear.html', {
        'form': {}, 'roles': roles, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('editar_usuarios')
def editar_usuario(request, id):
    usr   = get_object_or_404(Usuario.objects.select_related('rol'), pk=id)
    roles = Rol.objects.all()

    if request.method == 'POST':
        nombres          = request.POST.get('nombres', '').strip()
        apellidos        = request.POST.get('apellidos', '').strip()
        email            = request.POST.get('email', '').strip()
        password         = request.POST.get('password', '')
        rol_id           = request.POST.get('rol')
        estado           = request.POST.get('estado', 'ACTIVO')
        tipo_documento   = request.POST.get('tipo_documento', '')
        numero_documento = request.POST.get('numero_documento', '').strip()

        if not nombres or not email:
            messages.error(request, 'Nombre y email son obligatorios.')
            return render(request, 'admin/usuarios/editar.html', {
                'usuario_edit': usr, 'roles': roles,
                'form': request.POST, 'usuario': request.usuario
            })

        if Usuario.objects.filter(correo_usuario=email).exclude(pk=id).exists():
            messages.error(request, f'El email "{email}" ya está en uso.')
            return render(request, 'admin/usuarios/editar.html', {
                'usuario_edit': usr, 'roles': roles,
                'form': request.POST, 'usuario': request.usuario
            })

        usr.nombres_usuario   = nombres
        usr.apellidos_usuario = apellidos
        usr.correo_usuario    = email
        usr.estado_usuario    = estado
        usr.tipo_documento    = tipo_documento or None
        usr.numero_documento  = numero_documento or None
        if rol_id:
            usr.rol = get_object_or_404(Rol, pk=rol_id)
        if password:
            usr.set_password(password)
        usr.save()

        if str(usr.id) == str(request.session.get('usuario_id')) and estado == 'INACTIVO':
            request.session.flush()
            return redirect('login')

        messages.success(request, f'Usuario "{email}" actualizado.')
        return redirect('admin_usuarios')

    return render(request, 'admin/usuarios/editar.html', {
        'usuario_edit': usr, 'roles': roles, 'form': {}, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('eliminar_usuarios')
def eliminar_usuario(request, id):
    usr = get_object_or_404(Usuario, pk=id)
    if request.method == 'POST':
        email = usr.correo_usuario
        usr.delete()
        messages.success(request, f'Usuario "{email}" eliminado.')
        return redirect('admin_usuarios')
    return render(request, 'admin/usuarios/eliminar.html', {
        'usuario_edit': usr, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                         PROVEEDORES                                  ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_proveedores')
def listar_proveedores(request):
    busqueda    = request.GET.get('q', '').strip()
    proveedores = Proveedor.objects.order_by('nombre_proveedor')
    if busqueda:
        proveedores = proveedores.filter(
            Q(nombre_proveedor__icontains=busqueda) |
            Q(nit_proveedor__icontains=busqueda)
        )
    return render(request, 'admin/proveedores/lista.html', {
        'proveedores': proveedores, 'busqueda': busqueda, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('crear_proveedores')
def crear_proveedor(request):
    if request.method == 'POST':
        nit      = request.POST.get('nit_proveedor', '').strip()
        nombre   = request.POST.get('nombre_proveedor', '').strip()
        telefono = request.POST.get('telefono_proveedor', '').strip()
        correo   = request.POST.get('correo_proveedor', '').strip()

        if not nit or not nombre or not telefono or not correo:
            messages.error(request, 'Todos los campos son obligatorios.')
            return render(request, 'admin/proveedores/crear.html', {'form': request.POST, 'usuario': request.usuario})

        if Proveedor.objects.filter(nit_proveedor=nit).exists():
            messages.error(request, f'Ya existe un proveedor con NIT {nit}.')
            return render(request, 'admin/proveedores/crear.html', {'form': request.POST, 'usuario': request.usuario})

        try:
            Proveedor.objects.create(
                nit_proveedor=int(nit),
                nombre_proveedor=nombre,
                telefono_proveedor=int(telefono),
                correo_proveedor=correo,
            )
            messages.success(request, f'Proveedor "{nombre}" creado.')
            return redirect('admin_proveedores')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin/proveedores/crear.html', {'form': {}, 'usuario': request.usuario})

@admin_required
@permiso_requerido('editar_proveedores')
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, pk=id)

    if request.method == 'POST':
        # ── Limpiar AQUÍ, antes de todo ──
        nit      = request.POST.get('nit_proveedor', '').strip().replace('.', '').replace(',', '').replace(' ', '')
        nombre   = request.POST.get('nombre_proveedor', '').strip()
        telefono = request.POST.get('telefono_proveedor', '').strip().replace('.', '').replace(',', '').replace(' ', '')
        correo   = request.POST.get('correo_proveedor', '').strip()

        if not nit or not nombre or not telefono or not correo:
            messages.error(request, 'Todos los campos son obligatorios.')
            return render(request, 'admin/proveedores/editar.html', {
                'proveedor': proveedor, 'form': request.POST, 'usuario': request.usuario
            })

        # Ahora nit ya es limpio, int() no falla
        if Proveedor.objects.filter(nit_proveedor=int(nit)).exclude(pk=id).exists():
            messages.error(request, f'Ya existe otro proveedor con NIT {nit}.')
            return render(request, 'admin/proveedores/editar.html', {
                'proveedor': proveedor, 'form': request.POST, 'usuario': request.usuario
            })

        try:
            proveedor.nit_proveedor           = int(nit)
            proveedor.nombre_proveedor        = nombre
            proveedor.telefono_proveedor      = int(telefono)
            proveedor.correo_proveedor        = correo
            proveedor.contacto_proveedor      = request.POST.get('contacto_proveedor', '').strip() or None
            proveedor.direccion_proveedor     = request.POST.get('direccion_proveedor', '').strip() or None
            proveedor.observaciones_proveedor = request.POST.get('observaciones_proveedor', '').strip() or None
            proveedor.save()
            messages.success(request, f'Proveedor "{nombre}" actualizado.')
            return redirect('admin_proveedores')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin/proveedores/editar.html', {
        'proveedor': proveedor, 'form': {}, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('eliminar_proveedores')
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, pk=id)
    if request.method == 'POST':
        nombre = proveedor.nombre_proveedor
        proveedor.delete()
        messages.success(request, f'Proveedor "{nombre}" eliminado.')
        return redirect('admin_proveedores')
    return render(request, 'admin/proveedores/eliminar.html', {
        'proveedor': proveedor, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                          CATÁLOGOS                                   ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_catalogos')
def listar_catalogos(request):
    catalogos = Catalogo.objects.order_by('nombre_catalogo')
    return render(request, 'admin/catalogos/lista.html', {
        'catalogos': catalogos, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('crear_catalogos')
def crear_catalogo(request):
    if request.method == 'POST':
        nombre      = request.POST.get('nombre_catalogo', '').strip()
        descripcion = request.POST.get('descripcion_catalogo', '').strip()

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return render(request, 'admin/catalogos/crear.html', {
                'form': request.POST, 'usuario': request.usuario
            })

        Catalogo.objects.create(
            nombre_catalogo      = nombre,
            descripcion_catalogo = descripcion or None,
        )
        messages.success(request, f'Catálogo "{nombre}" creado.')
        return redirect('admin_catalogos')

    return render(request, 'admin/catalogos/crear.html', {
        'form': {}, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('crear_catalogos')
def editar_catalogo(request, id):
    catalogo = get_object_or_404(Catalogo, pk=id)
    if request.method == 'POST':
        nombre      = request.POST.get('nombre_catalogo', '').strip()
        descripcion = request.POST.get('descripcion_catalogo', '').strip()

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return render(request, 'admin/catalogos/editar.html', {
                'catalogo': catalogo, 'form': request.POST, 'usuario': request.usuario
            })

        catalogo.nombre_catalogo      = nombre
        catalogo.descripcion_catalogo = descripcion or None
        catalogo.save()
        messages.success(request, f'Catálogo "{nombre}" actualizado.')
        return redirect('admin_catalogos')

    return render(request, 'admin/catalogos/editar.html', {
        'catalogo': catalogo, 'form': {}, 'usuario': request.usuario
    })

@admin_required
@permiso_requerido('eliminar_catalogos')
def eliminar_catalogo(request, id):
    catalogo = get_object_or_404(Catalogo, pk=id)
    nombre = catalogo.nombre_catalogo
    catalogo.delete()
    messages.success(request, f'Catálogo "{nombre}" eliminado.')
    return redirect('admin_catalogos')

# ╔══════════════════════════════════════════════════════════════════════╗
# ║                         CATEGORÍAS                                   ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_categorias')
def listar_categorias(request):
    categorias = Categoria.objects.all().order_by('-id')
    return render(request, 'admin/categorias/lista.html', {
        'categorias': categorias, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('crear_categorias')
def crear_categoria(request):
    if request.method == 'POST':
        nombre      = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        estado      = request.POST.get('estado', 'ACTIVO')

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return render(request, 'admin/categorias/crear.html', {'form': request.POST, 'usuario': request.usuario})

        if Categoria.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, f'Ya existe una categoría "{nombre}".')
            return render(request, 'admin/categorias/crear.html', {'form': request.POST, 'usuario': request.usuario})

        Categoria.objects.create(nombre=nombre, descripcion=descripcion, estado=estado)
        messages.success(request, f'Categoría "{nombre}" creada.')
        return redirect('admin_categorias')

    return render(request, 'admin/categorias/crear.html', {'form': {}, 'usuario': request.usuario})


@admin_required
@permiso_requerido('editar_categorias')
def editar_categoria(request, id):
    categoria = get_object_or_404(Categoria, pk=id)

    if request.method == 'POST':
        nombre      = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        estado      = request.POST.get('estado', 'ACTIVO')

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return render(request, 'admin/categorias/editar.html', {
                'categoria': categoria, 'form': request.POST, 'usuario': request.usuario
            })

        if Categoria.objects.filter(nombre__iexact=nombre).exclude(pk=id).exists():
            messages.error(request, 'Ya existe otra categoría con ese nombre.')
            return render(request, 'admin/categorias/editar.html', {
                'categoria': categoria, 'form': request.POST, 'usuario': request.usuario
            })

        estado_anterior   = categoria.estado
        categoria.nombre  = nombre
        categoria.descripcion = descripcion
        categoria.estado  = estado
        categoria.save()

        if estado == 'INACTIVO' and estado_anterior == 'ACTIVO':
            Producto.objects.filter(categoria=categoria).update(estado='NO_DISPONIBLE')
        elif estado == 'ACTIVO' and estado_anterior == 'INACTIVO':
            Producto.objects.filter(categoria=categoria).update(estado='DISPONIBLE')

        messages.success(request, f'Categoría "{nombre}" actualizada.')
        return redirect('admin_categorias')

    return render(request, 'admin/categorias/editar.html', {
        'categoria': categoria, 'form': {}, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('eliminar_categorias')
def eliminar_categoria(request, id):
    categoria       = get_object_or_404(Categoria, pk=id)
    productos_count = Producto.objects.filter(categoria=categoria).count()

    if request.method == 'POST':
        tiene_historial = DetalleVenta.objects.filter(
            variacion__producto__categoria=categoria
        ).exists() or DetalleEntrada.objects.filter(
            variacion__producto__categoria=categoria
        ).exists()

        if tiene_historial:
            messages.error(request,
                f'No se puede eliminar "{categoria.nombre}" '
                f'porque tiene historial de ventas o entradas vinculado.'
            )
            return redirect('admin_categorias')

        nombre = categoria.nombre
        categoria.delete()
        messages.success(request, f'Categoría "{nombre}" y sus {productos_count} producto(s) eliminados.')
        return redirect('admin_categorias')

    return render(request, 'admin/categorias/eliminar.html', {
        'categoria': categoria, 'productos_count': productos_count, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                          PRODUCTOS                                   ║
# ╚══════════════════════════════════════════════════════════════════════╝

def _guardar_foto(archivo):
    if not archivo:
        return None
    tipos_ok = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp', 'image/gif']
    if archivo.content_type not in tipos_ok:
        raise ValueError('El archivo debe ser imagen (PNG, JPG, WEBP).')
    if archivo.size > 5 * 1024 * 1024:
        raise ValueError('La imagen no debe superar 5MB.')
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ext          = os.path.splitext(archivo.name)[1].lower()
    nombre_unico = f"{uuid.uuid4()}{ext}"
    with open(os.path.join(UPLOAD_DIR, nombre_unico), 'wb') as f:
        for chunk in archivo.chunks():
            f.write(chunk)
    return nombre_unico


def _eliminar_foto(nombre_foto):
    if nombre_foto:
        ruta = os.path.join(UPLOAD_DIR, nombre_foto)
        if os.path.exists(ruta):
            os.remove(ruta)


@admin_required
@permiso_requerido('ver_productos')
def listar_productos(request):
    busqueda   = request.GET.get('q', '').strip()
    categoria_id = request.GET.get('categoria', '')
    productos_qs = Producto.objects.select_related('categoria', 'proveedor', 'catalogo').order_by('-creado_en')
    categorias   = Categoria.objects.filter(estado='ACTIVO')

    if busqueda:
        productos_qs = productos_qs.filter(nombre_producto__icontains=busqueda)
    if categoria_id:
        productos_qs = productos_qs.filter(categoria_id=categoria_id)

    paginator = Paginator(productos_qs, 10)
    productos = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'admin/productos/lista.html', {
        'productos': productos, 'categorias': categorias,
        'busqueda': busqueda, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('crear_productos')
def crear_producto(request):
    categorias  = Categoria.objects.filter(estado='ACTIVO')
    proveedores = Proveedor.objects.all()
    catalogos   = Catalogo.objects.all()

    if request.method == 'POST':
        nombre       = request.POST.get('nombre_producto', '').strip()
        descripcion  = request.POST.get('descripcion_producto', '').strip()
        categoria_id = request.POST.get('categoria')
        proveedor_id = request.POST.get('proveedor')
        catalogo_id  = request.POST.get('catalogo')
        archivo      = request.FILES.get('archivoFoto')

        if not nombre or not categoria_id or not proveedor_id:
            messages.error(request, 'Nombre, categoría y proveedor son obligatorios.')
            return render(request, 'admin/productos/crear.html', {
                'form': request.POST, 'categorias': categorias,
                'proveedores': proveedores, 'catalogos': catalogos, 'usuario': request.usuario
            })

        try:
            foto = _guardar_foto(archivo)
            Producto.objects.create(
                nombre_producto=nombre,
                descripcion_producto=descripcion,
                categoria=get_object_or_404(Categoria, pk=categoria_id),
                proveedor=get_object_or_404(Proveedor, pk=proveedor_id),
                catalogo=Catalogo.objects.filter(pk=catalogo_id).first() if catalogo_id else None,
                foto=foto,
            )
            messages.success(request, f'Producto "{nombre}" creado.')
            return redirect('admin_productos')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin/productos/crear.html', {
        'form': {}, 'categorias': categorias,
        'proveedores': proveedores, 'catalogos': catalogos, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('editar_productos')
def editar_producto(request, id):
    producto    = get_object_or_404(Producto.objects.select_related('categoria', 'proveedor', 'catalogo'), pk=id)
    categorias  = Categoria.objects.filter(estado='ACTIVO')
    proveedores = Proveedor.objects.all()
    catalogos   = Catalogo.objects.all()

    if request.method == 'POST':
        nombre       = request.POST.get('nombre_producto', '').strip()
        descripcion  = request.POST.get('descripcion_producto', '').strip()
        categoria_id = request.POST.get('categoria')
        proveedor_id = request.POST.get('proveedor')
        catalogo_id  = request.POST.get('catalogo')
        estado       = request.POST.get('estado', producto.estado)
        archivo      = request.FILES.get('archivoFoto')

        if not nombre or not categoria_id or not proveedor_id:
            messages.error(request, 'Nombre, categoría y proveedor son obligatorios.')
            return render(request, 'admin/productos/editar.html', {
                'producto': producto, 'categorias': categorias,
                'proveedores': proveedores, 'catalogos': catalogos,
                'form': request.POST, 'usuario': request.usuario
            })

        try:
            if archivo:
                _eliminar_foto(producto.foto)
                producto.foto = _guardar_foto(archivo)
            producto.nombre_producto      = nombre
            producto.descripcion_producto = descripcion
            producto.categoria  = get_object_or_404(Categoria, pk=categoria_id)
            producto.proveedor  = get_object_or_404(Proveedor, pk=proveedor_id)
            producto.catalogo   = Catalogo.objects.filter(pk=catalogo_id).first() if catalogo_id else None
            producto.estado     = estado
            producto.save()
            messages.success(request, f'Producto "{nombre}" actualizado.')
            return redirect('admin_productos')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin/productos/editar.html', {
        'producto': producto, 'categorias': categorias,
        'proveedores': proveedores, 'catalogos': catalogos,
        'form': {}, 'usuario': request.usuario
    })
    



@admin_required
@permiso_requerido('eliminar_productos')
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, pk=id)

    if request.method == 'POST':
        tiene_ventas   = DetalleVenta.objects.filter(variacion__producto=producto).exists()
        tiene_entradas = DetalleEntrada.objects.filter(variacion__producto=producto).exists()

        if tiene_ventas or tiene_entradas:
            messages.error(request,
                f'No se puede eliminar "{producto.nombre_producto}" '
                f'porque tiene historial de ventas o entradas vinculado.'
            )
            return redirect('admin_productos')

        _eliminar_foto(producto.foto)
        nombre = producto.nombre_producto
        producto.delete()
        messages.success(request, f'Producto "{nombre}" eliminado.')
        return redirect('admin_productos')

    return render(request, 'admin/productos/eliminar.html', {
        'producto': producto, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                    VARIACIONES DE PRODUCTO                           ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_variaciones')
def listar_variaciones(request, producto_id):
    producto    = get_object_or_404(Producto, pk=producto_id)
    variaciones = VariacionProducto.objects.filter(producto=producto).order_by('talla', 'color')
    return render(request, 'admin/variaciones/lista.html', {
        'producto': producto, 'variaciones': variaciones, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('crear_variaciones')
def crear_variacion(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)

    if request.method == 'POST':
        talla          = request.POST.get('talla', '').strip().upper()
        color          = request.POST.get('color', '').strip()
        sku            = request.POST.get('sku_unico', '').strip()
        precio_compra  = request.POST.get('precio_compra', '')
        precio_base    = request.POST.get('precio_base', '')
        iva_porcentaje = request.POST.get('iva_porcentaje', '19')
        stock_actual   = request.POST.get('stock_actual', '0')
        stock_minimo   = request.POST.get('stock_minimo', '0')

        if not talla or not color or not sku or not precio_compra or not precio_base:
            messages.error(request, 'Todos los campos son obligatorios.')
            return render(request, 'admin/variaciones/crear.html', {
                'producto': producto, 'form': request.POST, 'usuario': request.usuario
            })

        if VariacionProducto.objects.filter(sku_unico=sku).exists():
            messages.error(request, f'Ya existe una variación con SKU "{sku}".')
            return render(request, 'admin/variaciones/crear.html', {
                'producto': producto, 'form': request.POST, 'usuario': request.usuario
            })

        try:
            VariacionProducto.objects.create(
                producto=producto,
                talla=talla,
                color=color,
                sku_unico=sku,
                precio_compra=Decimal(precio_compra),
                precio_base=float(precio_base),
                iva_porcentaje=float(iva_porcentaje),
                stock_actual=int(stock_actual),
                stock_minimo=int(stock_minimo),
            )
            producto.actualizar_estado()
            producto.save(update_fields=['estado', 'actualizado_en'])
            messages.success(request, f'Variación "{talla} - {color}" creada.')
            return redirect('admin_variaciones', producto_id=producto_id)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin/variaciones/crear.html', {
        'producto': producto, 'form': {}, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('editar_variaciones')
def editar_variacion(request, producto_id, id):
    producto  = get_object_or_404(Producto, pk=producto_id)
    variacion = get_object_or_404(VariacionProducto, pk=id, producto=producto)

    if request.method == 'POST':
        talla          = request.POST.get('talla', '').strip().upper()
        color          = request.POST.get('color', '').strip()
        sku            = request.POST.get('sku_unico', '').strip()
        precio_compra  = request.POST.get('precio_compra', '')
        precio_base    = request.POST.get('precio_base', '')
        iva_porcentaje = request.POST.get('iva_porcentaje', '19')
        stock_actual   = request.POST.get('stock_actual', '0')
        stock_minimo   = request.POST.get('stock_minimo', '0')

        if not talla or not color or not sku or not precio_compra or not precio_base:
            messages.error(request, 'Todos los campos son obligatorios.')
            return render(request, 'admin/variaciones/editar.html', {
                'producto': producto, 'variacion': variacion,
                'form': request.POST, 'usuario': request.usuario
            })

        if VariacionProducto.objects.filter(sku_unico=sku).exclude(pk=id).exists():
            messages.error(request, f'Ya existe otra variación con SKU "{sku}".')
            return render(request, 'admin/variaciones/editar.html', {
                'producto': producto, 'variacion': variacion,
                'form': request.POST, 'usuario': request.usuario
            })

        try:
            variacion.talla          = talla
            variacion.color          = color
            variacion.sku_unico      = sku
            variacion.precio_compra  = Decimal(precio_compra)
            variacion.precio_base    = float(precio_base)
            variacion.iva_porcentaje = float(iva_porcentaje)
            variacion.stock_actual   = int(stock_actual)
            variacion.stock_minimo   = int(stock_minimo)
            variacion.save()
            producto.actualizar_estado()
            producto.save(update_fields=['estado', 'actualizado_en'])
            messages.success(request, f'Variación "{talla} - {color}" actualizada.')
            return redirect('admin_variaciones', producto_id=producto_id)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin/variaciones/editar.html', {
        'producto': producto, 'variacion': variacion, 'form': {}, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('eliminar_variaciones')
def eliminar_variacion(request, producto_id, id):
    producto  = get_object_or_404(Producto, pk=producto_id)
    variacion = get_object_or_404(VariacionProducto, pk=id, producto=producto)

    if request.method == 'POST':
        tiene_ventas   = variacion.detalles_venta.exists()
        tiene_entradas = variacion.detalles_entrada.exists()

        if tiene_ventas or tiene_entradas:
            messages.error(request,
                f'No se puede eliminar "{variacion.talla} - {variacion.color}" '
                f'porque tiene historial vinculado. Pon el stock en 0 en su lugar.'
            )
            return redirect('admin_variaciones', producto_id=producto_id)

        label = f"{variacion.talla} - {variacion.color}"
        variacion.delete()
        producto.actualizar_estado()
        producto.save(update_fields=['estado', 'actualizado_en'])
        messages.success(request, f'Variación "{label}" eliminada.')
        return redirect('admin_variaciones', producto_id=producto_id)

    return render(request, 'admin/variaciones/eliminar.html', {
        'producto': producto, 'variacion': variacion, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                    INVENTARIO (ENTRADAS)                             ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_entradas')
def listar_entradas(request):
    busqueda = request.GET.get('q', '').strip()
    entradas = Entrada.objects.select_related('proveedor').order_by('-fecha_entrada')
    if busqueda:
        entradas = entradas.filter(proveedor__nombre_proveedor__icontains=busqueda)
    return render(request, 'admin/inventario/lista.html', {
        'entradas': entradas, 'busqueda': busqueda, 'usuario': request.usuario
    })

@admin_required
@permiso_requerido('crear_entradas')
def crear_entrada(request):
    proveedores = Proveedor.objects.all()
    variaciones = VariacionProducto.objects.select_related('producto').order_by('producto__nombre_producto', 'talla')

    if request.method == 'POST':
        proveedor_id  = request.POST.get('proveedor')
        fecha_entrada = request.POST.get('fecha_entrada')
        total_filas   = int(request.POST.get('totalFilas', 0))

        variacion_ids = []
        cantidades    = []
        precios       = []

        for i in range(total_filas):
            vid    = request.POST.get(f'variacion_{i}')
            cant   = request.POST.get(f'cantidad_{i}')
            precio = request.POST.get(f'precio_{i}')
            if vid and cant and precio:
                variacion_ids.append(vid)
                cantidades.append(cant)
                precios.append(precio)

        if not proveedor_id or not fecha_entrada or not variacion_ids:
            messages.error(request, 'Proveedor, fecha y al menos una variación son obligatorios.')
            return render(request, 'admin/inventario/crear.html', {
                'proveedores': proveedores, 'variaciones': variaciones,
                'form': request.POST, 'usuario': request.usuario
            })

        try:
            with transaction.atomic():
                entrada = Entrada.objects.create(
                 proveedor=get_object_or_404(Proveedor, pk=proveedor_id),
                fecha_entrada=fecha_entrada,
                total_entrada=0,
                )
                total = 0
                for vid, cant, precio in zip(variacion_ids, cantidades, precios):
                    if not cant or int(cant) <= 0:
                        continue
                    detalle = DetalleEntrada.objects.create(
                    entrada=entrada,
                    variacion=get_object_or_404(VariacionProducto, pk=vid),
                    cantidades=int(cant),
                    precio_comprado=int(precio),
                    )
                    detalle.aplicar_stock()
                    total += int(cant)

                entrada.total_entrada = total
                entrada.save(update_fields=['total_entrada'])

            messages.success(request, f'Entrada de inventario #{entrada.id} registrada.')
            return redirect('admin_inventario')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin/inventario/crear.html', {
        'proveedores': proveedores, 'variaciones': variaciones,
        'form': {}, 'usuario': request.usuario
    })

@admin_required
def detalle_entrada(request, id):
    entrada  = get_object_or_404( Entrada.objects.select_related('proveedor'), pk=id)
    detalles = DetalleEntrada.objects.select_related('variacion__producto').filter(entrada=entrada)
    return render(request, 'admin/inventario/detalle.html', {
        'entrada':entrada, 'detalles': detalles, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('eliminar_entradas')
def eliminar_entrada(request, id):
    entrada = get_object_or_404(Entrada, pk=id)
    if request.method == 'POST':
        entrada.delete()
        messages.success(request, f'Entrada #{id} eliminada.')
        return redirect('admin_inventario')
    return render(request, 'admin/inventario/eliminar.html', {
        'entrada': entrada, 'usuario': request.usuario
    })


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                           VENTAS                                     ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_ventas')
def listar_ventas(request):
    busqueda_usuario  = request.GET.get('busqueda_usuario', '').strip()
    busqueda_producto = request.GET.get('busqueda_producto', '').strip()
    filtro_fecha      = request.GET.get('filtro', '')
    fecha_inicio      = request.GET.get('fecha_inicio', '')
    fecha_fin         = request.GET.get('fecha_fin', '')

    ventas = Venta.objects.select_related('usuario').prefetch_related(
        'detalles__variacion__producto'
    ).order_by('-id')

    filtro_label = 'Todas las ventas'

    if filtro_fecha == 'hoy':
        ventas = ventas.filter(fecha__date=timezone.now().date())
        filtro_label = 'Ventas del día'
    elif filtro_fecha == 'mes':
        hoy    = timezone.now()
        ventas = ventas.filter(fecha__year=hoy.year, fecha__month=hoy.month)
        filtro_label = 'Ventas del mes'
    elif fecha_inicio and fecha_fin:
        ventas = ventas.filter(fecha__date__gte=fecha_inicio, fecha__date__lte=fecha_fin)
        filtro_label = f'Del {fecha_inicio} al {fecha_fin}'
    elif busqueda_usuario:
        ventas = ventas.filter(
            Q(usuario__nombres_usuario__icontains=busqueda_usuario) |
            Q(usuario__apellidos_usuario__icontains=busqueda_usuario) |
            Q(usuario__correo_usuario__icontains=busqueda_usuario)
        )
        filtro_label = f'Cliente: {busqueda_usuario}'
    elif busqueda_producto:
        ventas = ventas.filter(
            detalles__variacion__producto__nombre_producto__icontains=busqueda_producto
        ).distinct()
        filtro_label = f'Producto: {busqueda_producto}'

    return render(request, 'admin/ventas/lista.html', {
        'ventas':            ventas,
        'filtro_label':      filtro_label,
        'busqueda_usuario':  busqueda_usuario,
        'busqueda_producto': busqueda_producto,
        'fecha_inicio':      fecha_inicio,
        'fecha_fin':         fecha_fin,
        'usuario':           request.usuario,
    })


@admin_required
def detalle_venta(request, id):
    venta    = get_object_or_404(Venta.objects.select_related('usuario'), pk=id)
    detalles = DetalleVenta.objects.select_related('variacion__producto').filter(venta=venta)
    envio    = Envio.objects.filter(venta=venta).first()
    return render(request, 'admin/ventas/detalle.html', {
        'venta': venta, 'detalles': detalles, 'envio': envio, 'usuario': request.usuario
    })


@admin_required
def exportar_ventas_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'Ejecuta: pip install openpyxl')
        return redirect('admin_ventas')

    ventas = Venta.objects.select_related('usuario').prefetch_related(
        'detalles__variacion__producto'
    ).order_by('-fecha')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas TiendaMoa'

    headers     = ['ID Venta', 'Fecha', 'Cliente', 'Email', 'Producto', 'Talla', 'Color', 'SKU', 'Cantidad', 'Precio Unit.', 'Subtotal', 'Total Venta']
    header_fill = PatternFill(start_color='1447A8', end_color='1447A8', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    for venta in ventas:
        for d in venta.detalles.all():
            ws.append([
                venta.id,
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                venta.usuario.get_full_name(),
                venta.usuario.correo_usuario,
                d.variacion.producto.nombre_producto,
                d.variacion.talla,
                d.variacion.color,
                d.variacion.sku_unico,
                d.cantidad,
                d.precio_unitario,
                d.subtotal,
                venta.monto_final,
            ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ventas_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response


@admin_required
def exportar_ventas_pdf(request):
    try:
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        messages.error(request, 'Ejecuta: pip install reportlab')
        return redirect('admin_ventas')

    NARANJA  = colors.HexColor('#FF6B35')
    NAVY     = colors.HexColor('#1e2d4a')
    BLANCO   = colors.white
    GRIS_CLR = colors.HexColor('#F5F4F0')
    GRIS2    = colors.HexColor('#e8e4de')
    MUTED    = colors.HexColor('#7a8fa6')
    DARK     = colors.HexColor('#1A1A2E')

    ventas = Venta.objects.select_related('usuario').prefetch_related(
        'detalles__variacion__producto'
    ).order_by('id')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter,
                                leftMargin=28*mm, rightMargin=28*mm,
                                topMargin=20*mm, bottomMargin=20*mm)

    estilo_titulo = ParagraphStyle('titulo', fontName='Helvetica-Bold', fontSize=20, textColor=DARK, spaceAfter=2)
    estilo_sub    = ParagraphStyle('sub',    fontName='Helvetica',      fontSize=9,  textColor=MUTED, spaceAfter=10)
    estilo_pie    = ParagraphStyle('pie',    fontName='Helvetica-Oblique', fontSize=8, textColor=MUTED, alignment=TA_CENTER)

    now_str        = timezone.now().strftime('%d/%m/%Y %H:%M')
    total_ventas   = ventas.count()
    total_ingresos = sum(v.monto_final for v in ventas)
    page_w         = letter[0] - 56*mm
    elements       = []

    banner = Table([['TiendaMoa07']], colWidths=[page_w])
    banner.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), NARANJA),
        ('TEXTCOLOR',     (0, 0), (-1, 0), BLANCO),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 11),
        ('ALIGN',         (0, 0), (-1, 0), 'LEFT'),
        ('TOPPADDING',    (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 9),
        ('LEFTPADDING',   (0, 0), (-1, 0), 12),
    ]))
    elements.append(banner)
    elements.append(Spacer(1, 10))
    elements.append(Paragraph('Reporte de Ventas', estilo_titulo))
    elements.append(Paragraph(f'Generado el {now_str}  ·  Colombia, Bogotá D.C.', estilo_sub))
    elements.append(HRFlowable(width='100%', thickness=3, color=NARANJA, spaceBefore=8, spaceAfter=10))

    resumen = Table([[
        f'{total_ventas}\nVentas totales',
        f'${total_ingresos:,.0f}\nIngresos totales (COP)',
    ]], colWidths=[page_w / 2] * 2)
    resumen.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), GRIS_CLR),
        ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 11),
        ('TEXTCOLOR',     (0, 0), (-1, -1), DARK),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING',    (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(resumen)
    elements.append(Spacer(1, 14))

    data = [['#', 'Fecha', 'Cliente', 'Producto', 'Talla', 'Cant.', 'Total']]
    fila = 1
    for venta in ventas:
        for d in venta.detalles.all():
            data.append([
                str(fila),
                venta.fecha.strftime('%d/%m/%Y'),
                venta.usuario.get_full_name(),
                d.variacion.producto.nombre_producto,
                d.variacion.talla,
                str(d.cantidad),
                f'${venta.monto_final:,.0f}',
            ])
            fila += 1

    col_widths = [12*mm, 22*mm, 40*mm, 40*mm, 14*mm, 14*mm, 24*mm]
    table      = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR',     (0, 0), (-1, 0), BLANCO),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('TEXTCOLOR',     (0, 1), (-1, -1), DARK),
        ('TOPPADDING',    (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LINEBELOW',     (0, 0), (-1, -2), 0.3, GRIS2),
        ('BOX',           (0, 0), (-1, -1), 0.5, GRIS2),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))
    elements.append(HRFlowable(width='100%', thickness=0.5, color=GRIS2, spaceBefore=4, spaceAfter=8))
    elements.append(Paragraph('© 2026 TiendaMoa07  ·  Colombia, Bogotá D.C.', estilo_pie))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ventas_{timezone.now().strftime("%d%m%Y")}.pdf"'
    return response


# ╔══════════════════════════════════════════════════════════════════════╗
# ║                           ENVÍOS                                     ║
# ╚══════════════════════════════════════════════════════════════════════╝

@admin_required
@permiso_requerido('ver_envios')
def listar_envios(request):
    busqueda = request.GET.get('q', '').strip()
    envios   = Envio.objects.select_related('venta__usuario', 'usuario').order_by('-id')
    if busqueda:
        envios = envios.filter(
            Q(numero_guia__icontains=busqueda)       |
            Q(empresa_transportadora__icontains=busqueda) |
            Q(estado_envio__icontains=busqueda)      |
            Q(venta__usuario__nombres_usuario__icontains=busqueda)
        )
    return render(request, 'admin/envios/lista.html', {
        'envios': envios, 'busqueda': busqueda, 'usuario': request.usuario
    })


@admin_required
def detalle_envio(request, id):
    envio    = get_object_or_404(Envio.objects.select_related('venta__usuario', 'usuario'), pk=id)
    detalles = DetalleVenta.objects.select_related('variacion__producto').filter(venta=envio.venta)
    return render(request, 'admin/envios/detalle.html', {
        'envio': envio, 'detalles': detalles, 'usuario': request.usuario
    })


@admin_required
@permiso_requerido('editar_envios')
def editar_envio(request, id):

    envio = get_object_or_404(
        Envio.objects.select_related('venta__usuario'),
        pk=id
    )

    if request.method == 'POST':

        departamento     = request.POST.get('departamento_envio', '').strip()
        ciudad            = request.POST.get('ciudad_envio', '').strip()
        barrio            = request.POST.get('barrio_envio', '').strip()
        direccion         = request.POST.get('direccion_envio', '').strip()
        tipo_vivienda     = request.POST.get('tipo_vivienda', envio.tipo_vivienda)
        especificaciones  = request.POST.get('especificaciones_llegada', '').strip()
        telefono          = request.POST.get('telefono_llegada', '').strip()

        empresa           = request.POST.get('empresa_transportadora', '').strip()
        numero_guia       = request.POST.get('numero_guia', '').strip()
        estado            = request.POST.get('estado_envio', envio.estado_envio)

        fecha_envio       = request.POST.get('fecha_envio')
        fecha_llegada     = request.POST.get('fecha_estipulada_llegada')

        costo             = request.POST.get('costo_envio', '0')

        # SOLO validar si vienen esos campos
        if 'direccion_envio' in request.POST:

            if not departamento or not ciudad or not direccion or not telefono:

                messages.error(
                    request,
                    'Departamento, ciudad, dirección y teléfono son obligatorios.'
                )

                return render(request, 'admin/envios/editar.html', {
                    'envio': envio,
                    'form': request.POST,
                    'departamentos': sorted(CIUDADES_POR_DEPARTAMENTO.keys()),
                    'ciudades_por_departamento': CIUDADES_POR_DEPARTAMENTO,
                    'usuario': request.usuario
                })

        if numero_guia and Envio.objects.filter(
            numero_guia=numero_guia
        ).exclude(pk=id).exists():

            messages.error(
                request,
                f'Ya existe otro envío con número de guía "{numero_guia}".'
            )

            return render(request, 'admin/envios/editar.html', {
                'envio': envio,
                'form': request.POST,
                'departamentos': sorted(CIUDADES_POR_DEPARTAMENTO.keys()),
                'ciudades_por_departamento': CIUDADES_POR_DEPARTAMENTO,
                'usuario': request.usuario
            })

        envio.departamento_envio       = departamento or envio.departamento_envio
        envio.ciudad_envio             = ciudad or envio.ciudad_envio
        envio.barrio_envio             = barrio or envio.barrio_envio
        envio.direccion_envio          = direccion or envio.direccion_envio
        envio.tipo_vivienda            = tipo_vivienda

        envio.especificaciones_llegada = especificaciones or envio.especificaciones_llegada
        envio.telefono_llegada         = telefono or envio.telefono_llegada

        envio.empresa_transportadora   = empresa or envio.empresa_transportadora
        envio.numero_guia              = numero_guia or envio.numero_guia

        envio.estado_envio             = estado

        envio.fecha_envio = fecha_envio or envio.fecha_envio
        envio.fecha_estipulada_llegada = (
            fecha_llegada or envio.fecha_estipulada_llegada
        )

        envio.costo_envio = Decimal(costo)

        envio.save()

        messages.success(
            request,
            f'Envío #{id} actualizado correctamente.'
        )

        return redirect('admin_envios')

    return render(request, 'admin/envios/editar.html', {
        'envio': envio,
        'form': {},
        'departamentos': sorted(CIUDADES_POR_DEPARTAMENTO.keys()),
        'ciudades_por_departamento': CIUDADES_POR_DEPARTAMENTO,
        'usuario': request.usuario
    })

@admin_required
@permiso_requerido('eliminar_envios')
def eliminar_envio(request, id):
    envio = get_object_or_404(Envio, pk=id)
    if request.method == 'POST':
        envio.delete()
        messages.success(request, f'Envío #{id} eliminado.')
        return redirect('admin_envios')
    return render(request, 'admin/envios/eliminar.html', {
        'envio': envio, 'usuario': request.usuario
    })


@admin_required
def envio_cambiar_estado(request, id):
    if request.method == 'POST':
        envio = get_object_or_404(Envio, pk=id)
        envio.estado_envio = request.POST.get('estado') or request.POST.get('estado_envio') or envio.estado_envio
        envio.save()
        messages.success(request, f'Estado del envío #{id} actualizado a "{envio.get_estado_envio_display()}".')
    return redirect('admin_envios')

@admin_required
def reporte_inventario(request):
    """Reporte 1 — Inventario por Producto"""
    filtro_talla      = request.GET.get('talla', '')
    filtro_color      = request.GET.get('color', '')
    filtro_categoria  = request.GET.get('categoria', '')
    filtro_stock      = request.GET.get('estado_stock', '')
    exportar          = request.GET.get('export', '')

    variaciones_qs = VariacionProducto.objects.select_related(
        'producto__categoria', 'producto__proveedor'
    ).order_by('producto__nombre_producto', 'talla', 'color')

    if filtro_talla:
        variaciones_qs = variaciones_qs.filter(talla__icontains=filtro_talla)
    if filtro_color:
        variaciones_qs = variaciones_qs.filter(color__icontains=filtro_color)
    if filtro_categoria:
        variaciones_qs = variaciones_qs.filter(producto__categoria_id=filtro_categoria)
    if filtro_stock == 'bajo_stock':
        variaciones_qs = [v for v in variaciones_qs if v.bajo_stock and not v.sin_stock]
    elif filtro_stock == 'agotado':
        variaciones_qs = [v for v in variaciones_qs if v.sin_stock]
    elif filtro_stock == 'disponible':
        variaciones_qs = [v for v in variaciones_qs if not v.sin_stock]
    else:
        variaciones_qs = list(variaciones_qs)

    # Estadísticas globales
    todas             = VariacionProducto.objects.select_related('producto__categoria')
    total_variaciones = todas.count()
    total_agotados    = sum(1 for v in todas if v.sin_stock)
    total_bajo_stock  = sum(1 for v in todas if v.bajo_stock and not v.sin_stock)
    stock_total       = todas.aggregate(t=Sum('stock_actual'))['t'] or 0
    total_productos   = Producto.objects.count()

    # Rotación
    top_vendidos = (
        DetalleVenta.objects
        .values('variacion__producto__nombre_producto')
        .annotate(unidades_vendidas=Sum('cantidad'))
        .order_by('-unidades_vendidas')
    )
    productos_mayor_rotacion = [
        {'nombre': x['variacion__producto__nombre_producto'], 'unidades_vendidas': x['unidades_vendidas']}
        for x in top_vendidos[:5]
    ]
    productos_menor_rotacion = [
        {'nombre': x['variacion__producto__nombre_producto'], 'unidades_vendidas': x['unidades_vendidas']}
        for x in top_vendidos.order_by('unidades_vendidas')[:5]
    ]
    total_mayor_rotacion = productos_mayor_rotacion[0]['unidades_vendidas'] if productos_mayor_rotacion else 0

    # Filtros disponibles
    tallas_disponibles     = VariacionProducto.objects.values_list('talla', flat=True).distinct().order_by('talla')
    colores_disponibles    = VariacionProducto.objects.values_list('color', flat=True).distinct().order_by('color')
    categorias_disponibles = Categoria.objects.filter(estado='ACTIVO')

    # Exportar Excel
    if exportar == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Inventario'
            headers = ['Producto', 'Categoría', 'Talla', 'Color', 'SKU', 'Stock', 'Stock Mínimo', 'Estado']
            header_fill = PatternFill('solid', fgColor='144272')
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for row, v in enumerate(variaciones_qs, 2):
                if v.sin_stock:
                    estado = 'Agotado'
                elif v.bajo_stock:
                    estado = 'Bajo stock'
                else:
                    estado = 'Disponible'
                ws.append([
                    v.producto.nombre_producto,
                    v.producto.categoria.nombre if v.producto.categoria else '—',
                    v.talla,
                    v.color,
                    v.sku_unico,
                    v.stock_actual,
                    v.stock_minimo,
                    estado,
                ])
            from django.http import HttpResponse
            import io
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
            return response
        except ImportError:
            messages.error(request, 'Ejecuta: pip install openpyxl')

    return render(request, 'admin/reportes/inventario.html', {
        'variaciones':               variaciones_qs,
        'total_variaciones':         total_variaciones,
        'total_agotados':            total_agotados,
        'total_bajo_stock':          total_bajo_stock,
        'stock_total':               stock_total,
        'total_productos':           total_productos,
        'total_mayor_rotacion':      total_mayor_rotacion,
        'productos_mayor_rotacion':  productos_mayor_rotacion,
        'productos_menor_rotacion':  productos_menor_rotacion,
        'tallas_disponibles':        tallas_disponibles,
        'colores_disponibles':       colores_disponibles,
        'categorias_disponibles':    categorias_disponibles,
        'marcas_disponibles':        [],
        'ciudades_disponibles':      [],
        'departamentos_disponibles': [],
        'usuario':                   request.usuario,
    })
 
@admin_required
def reporte_ventas(request):
    """Reporte 2 — Ventas por Fecha y Usuario"""
    busqueda_usuario  = request.GET.get('busqueda_usuario', '').strip()
    busqueda_producto = request.GET.get('busqueda_producto', '').strip()
    filtro_fecha      = request.GET.get('filtro', '')
    fecha_inicio      = request.GET.get('fecha_inicio', '')
    fecha_fin         = request.GET.get('fecha_fin', '')
 
    ventas = Venta.objects.select_related('usuario').prefetch_related(
        'detalles__variacion__producto'
    ).order_by('-fecha')
 
    filtro_label = 'Todas las ventas'
 
    if filtro_fecha == 'hoy':
        ventas = ventas.filter(fecha__date=timezone.now().date())
        filtro_label = 'Ventas del día'
    elif filtro_fecha == 'mes':
        hoy    = timezone.now()
        ventas = ventas.filter(fecha__year=hoy.year, fecha__month=hoy.month)
        filtro_label = 'Ventas del mes'
    elif filtro_fecha == 'semana':
        desde = timezone.now().date() - timedelta(days=7)
        ventas = ventas.filter(fecha__date__gte=desde)
        filtro_label = 'Últimos 7 días'
    elif fecha_inicio and fecha_fin:
        ventas = ventas.filter(fecha__date__gte=fecha_inicio, fecha__date__lte=fecha_fin)
        filtro_label = f'Del {fecha_inicio} al {fecha_fin}'
    elif busqueda_usuario:
        ventas = ventas.filter(
            Q(usuario__nombres_usuario__icontains=busqueda_usuario) |
            Q(usuario__apellidos_usuario__icontains=busqueda_usuario) |
            Q(usuario__correo_usuario__icontains=busqueda_usuario)
        )
        filtro_label = f'Cliente: {busqueda_usuario}'
    elif busqueda_producto:
        ventas = ventas.filter(
            detalles__variacion__producto__nombre_producto__icontains=busqueda_producto
        ).distinct()
        filtro_label = f'Producto: {busqueda_producto}'
 
    ventas_list = list(ventas)
 
    # Estadísticas del filtro actual
    total_ingresos  = sum(v.monto_final for v in ventas_list)
    total_productos = sum(v.cantidad_productos for v in ventas_list)
 
    # Top vendedores
    top_usuarios = (
        Venta.objects.values(
            'usuario__nombres_usuario', 'usuario__apellidos_usuario', 'usuario__correo_usuario'
        )
        .annotate(num_ventas=Count('id'), total_gastado=Sum('monto_final'))
        .order_by('-num_ventas')[:5]
    )
 
    # Ventas por día (últimos 30 días)
    desde_30 = timezone.now().date() - timedelta(days=30)
    ventas_por_dia = (
        Venta.objects
        .filter(fecha__date__gte=desde_30)
        .annotate(dia=TruncDay('fecha'))
        .values('dia')
        .annotate(total=Sum('monto_final'), cantidad=Count('id'))
        .order_by('dia')
    )
 
    return render(request, 'admin/reportes/ventas.html', {
        'ventas':             ventas_list,
        'filtro_label':       filtro_label,
        'busqueda_usuario':   busqueda_usuario,
        'busqueda_producto':  busqueda_producto,
        'fecha_inicio':       fecha_inicio,
        'fecha_fin':          fecha_fin,
        'total_ingresos':     total_ingresos,
        'total_productos':    total_productos,
        'top_usuarios':       top_usuarios,
        'ventas_por_dia':     list(ventas_por_dia),
        'usuario':            request.usuario,
    })
 
 
@admin_required
def reporte_categorias(request):
    """Reporte 3 — Ventas por Categoría"""
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin    = request.GET.get('fecha_fin', '')
 
    detalles_qs = DetalleVenta.objects.select_related(
        'variacion__producto__categoria', 'venta'
    )
    if fecha_inicio and fecha_fin:
        detalles_qs = detalles_qs.filter(
            venta__fecha__date__gte=fecha_inicio,
            venta__fecha__date__lte=fecha_fin
        )
 
    # Ventas agrupadas por categoría
    por_categoria = (
        detalles_qs
        .values('variacion__producto__categoria__nombre')
        .annotate(
            total_unidades=Sum('cantidad'),
            total_ingresos=Sum('subtotal'),   # subtotal es property, usamos precio*cant
            num_ventas=Count('venta', distinct=True),
        )
        .order_by('-total_ingresos')
    )
    # subtotal es @property, no un campo DB → calculamos con precio_unitario*cantidad
    por_categoria = (
        detalles_qs
        .values('variacion__producto__categoria__nombre')
        .annotate(
            total_unidades=Sum('cantidad'),
            total_ingresos=Sum(
                models_producto_ingreso()
            ),
            num_ventas=Count('venta', distinct=True),
        )
        .order_by('-total_ingresos')
    )
 
    # Usamos annotate con expresión directa
    from django.db.models import F, ExpressionWrapper, FloatField
    por_categoria = (
        detalles_qs
        .annotate(ingreso_item=ExpressionWrapper(
            F('cantidad') * F('precio_unitario'), output_field=FloatField()
        ))
        .values('variacion__producto__categoria__nombre')
        .annotate(
            total_unidades=Sum('cantidad'),
            total_ingresos=Sum('ingreso_item'),
            num_ventas=Count('venta', distinct=True),
        )
        .order_by('-total_ingresos')
    )
 
    total_global = sum(c['total_ingresos'] or 0 for c in por_categoria)
 
    # Producto más vendido por categoría
    top_por_cat = (
        DetalleVenta.objects
        .annotate(ingreso_item=ExpressionWrapper(
            F('cantidad') * F('precio_unitario'), output_field=FloatField()
        ))
        .values(
            'variacion__producto__categoria__nombre',
            'variacion__producto__nombre_producto'
        )
        .annotate(total_und=Sum('cantidad'), total_ing=Sum('ingreso_item'))
        .order_by('variacion__producto__categoria__nombre', '-total_und')
    )
 
    return render(request, 'admin/reportes/categorias.html', {
        'por_categoria': list(por_categoria),
        'total_global':  total_global,
        'top_por_cat':   list(top_por_cat),
        'fecha_inicio':  fecha_inicio,
        'fecha_fin':     fecha_fin,
        'usuario':       request.usuario,
    })
 
 
def models_producto_ingreso():
    """Placeholder — no se usa, la expresión se construye inline."""
    return None
 
 
@admin_required
def reporte_envios(request):
    """Reporte 4 — Estado de Envíos"""
    filtro_estado  = request.GET.get('estado', '')
    filtro_empresa = request.GET.get('empresa', '')
    busqueda       = request.GET.get('q', '').strip()
    fecha_inicio   = request.GET.get('fecha_inicio', '')
    fecha_fin      = request.GET.get('fecha_fin', '')
 
    envios = Envio.objects.select_related('venta__usuario', 'usuario').order_by('-id')
 
    if filtro_estado:
        envios = envios.filter(estado_envio=filtro_estado)
    if filtro_empresa:
        envios = envios.filter(empresa_transportadora__icontains=filtro_empresa)
    if busqueda:
        envios = envios.filter(
            Q(ciudad_envio__icontains=busqueda) |
            Q(departamento_envio__icontains=busqueda) |
            Q(venta__usuario__nombres_usuario__icontains=busqueda) |
            Q(numero_guia__icontains=busqueda)
        )
    if fecha_inicio and fecha_fin:
        envios = envios.filter(fecha_envio__date__gte=fecha_inicio, fecha_envio__date__lte=fecha_fin)
 
    # Estadísticas globales (siempre sobre todos)
    todos = Envio.objects.all()
    stats = {e[0]: 0 for e in Envio.ESTADO_CHOICES}
    for e in todos:
        stats[e.estado_envio] = stats.get(e.estado_envio, 0) + 1
 
    # Top empresas transportadoras
    top_empresas = (
        Envio.objects
        .values('empresa_transportadora')
        .annotate(total=Count('id'))
        .order_by('-total')[:6]
    )
 
    # Top destinos
    top_destinos = (
        Envio.objects
        .values('ciudad_envio', 'departamento_envio')
        .annotate(total=Count('id'))
        .order_by('-total')[:8]
    )
 
    empresas_disponibles = (
        Envio.objects
        .values_list('empresa_transportadora', flat=True)
        .distinct()
        .order_by('empresa_transportadora')
    )
 
    return render(request, 'admin/reportes/envios.html', {
        'envios':               envios,
        'stats':                stats,
        'top_empresas':         top_empresas,
        'top_destinos':         top_destinos,
        'filtro_estado':        filtro_estado,
        'filtro_empresa':       filtro_empresa,
        'busqueda':             busqueda,
        'fecha_inicio':         fecha_inicio,
        'fecha_fin':            fecha_fin,
        'estados':              Envio.ESTADO_CHOICES,
        'empresas_disponibles': empresas_disponibles,
        'usuario':              request.usuario,
    })
 
 
@admin_required
def reporte_clientes(request):
    """Reporte 5 — Comportamiento de Clientes"""
    busqueda     = request.GET.get('q', '').strip()
    filtro_orden = request.GET.get('orden', 'ventas')

    clientes_qs = Usuario.objects.filter(
        rol__nombre_rol__iexact='CLIENTE'
    ).prefetch_related('ventas__detalles')

    if busqueda:
        clientes_qs = clientes_qs.filter(
            Q(nombres_usuario__icontains=busqueda) |
            Q(apellidos_usuario__icontains=busqueda) |
            Q(correo_usuario__icontains=busqueda)
        )

    clientes_qs = clientes_qs.annotate(
        num_ventas=Count('ventas'),
        total_gastado=Sum('ventas__monto_final'),
    )

    if filtro_orden == 'gastado':
        clientes_qs = clientes_qs.order_by('-total_gastado')
    else:
        clientes_qs = clientes_qs.order_by('-num_ventas')

    clientes_list = list(clientes_qs)

    total_clientes = Usuario.objects.filter(rol__nombre_rol__iexact='CLIENTE').count()
    clientes_activos = Venta.objects.values('usuario_id').distinct().count()

    top_clientes = sorted(
        [c for c in clientes_list if c.total_gastado],
        key=lambda c: c.total_gastado or 0,
        reverse=True
    )[:5]

    return render(request, 'admin/reportes/clientes.html', {
        'clientes':           clientes_list,
        'busqueda':           busqueda,
        'filtro_orden':       filtro_orden,
        'total_clientes':     total_clientes,
        'clientes_activos':   clientes_activos,
        'clientes_inactivos': total_clientes - clientes_activos,
        'top_clientes':       top_clientes,
        'usuario':            request.usuario,
    })


@admin_required
def reportes_hub(request):
    return render(request, 'admin/reportes/reportes_hub.html')

    


# ══════════════════════════════════════════════════════
# RESTABLECIMIENTO DE CONTRASEÑA 1
# ══════════════════════════════════════════════════════
def solicitar_reset(request):
    if request.method == 'POST':
        correo = request.POST.get('correo', '').strip()
        try:
            usuario = Usuario.objects.get(correo_usuario=correo)
            TokenReset.objects.filter(usuario=usuario, usado=False).update(usado=True)
            token = TokenReset.objects.create(usuario=usuario)
            enlace = request.build_absolute_uri(f'/reset-password/{token.token}/')
            resend.Emails.send({
                "from": "TiendaMoa <onboarding@resend.dev>",
                "to": usuario.correo_usuario,
                "subject": "Restablecer contraseña - TiendaMoa",
                "html": f"""
                <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                    <h2 style="color: #144272;">🏷️ TiendaMoa</h2>
                    <p>Hola <strong>{usuario.nombres_usuario}</strong>,</p>
                    <p>Recibimos una solicitud para restablecer tu contraseña.</p>
                    <a href="{enlace}" style="
                        display: inline-block;
                        padding: 14px 28px;
                        background: linear-gradient(135deg, #667eea, #764ba2);
                        color: white;
                        text-decoration: none;
                        border-radius: 10px;
                        font-weight: bold;
                        margin: 20px 0;
                    ">Restablecer contraseña</a>
                    <p style="color: #888; font-size: 13px;">Este enlace expira en 24 horas. Si no solicitaste esto, ignora este correo.</p>
                </div>
                """,
            })
        except Usuario.DoesNotExist:
            pass
        except Exception as e:
            print(f"ERROR RESEND: {e}")
        return redirect('reset_enviado')
    return render(request, 'registration/password_reset_form.html')


def reset_enviado(request):
    return render(request, 'registration/password_reset_done.html')


def confirmar_reset(request, token):
    try:
        token_obj = TokenReset.objects.get(token=token)
    except TokenReset.DoesNotExist:
        return render(request, 'registration/password_reset_confirm.html', {'validlink': False})

    if not token_obj.esta_vigente():
        return render(request, 'registration/password_reset_confirm.html', {'validlink': False})

    error = None
    if request.method == 'POST':
        nueva = request.POST.get('new_password1', '')
        confirmar = request.POST.get('new_password2', '')
        if nueva != confirmar:
            error = 'Las contraseñas no coinciden.'
        elif len(nueva) < 6:
            error = 'La contraseña debe tener al menos 6 caracteres.'
        else:
            token_obj.usuario.set_password(nueva)
            token_obj.usuario.save()
            token_obj.usado = True
            token_obj.save()
            return redirect('reset_completo')

    return render(request, 'registration/password_reset_confirm.html', {'validlink': True, 'error': error})


def reset_completo(request):
    return render(request, 'registration/password_reset_complete.html')


